"""
Excel Parser — Extract text from .xlsx / .xls spreadsheets.
Uses openpyxl to read all sheets and convert to structured text.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}


def can_parse(file_type: str) -> bool:
    return file_type.lower().lstrip(".") in {ext.lstrip(".") for ext in SUPPORTED_EXTENSIONS}


async def parse(file_path: str) -> dict[str, Any]:
    """
    Parse an Excel file. Converts each sheet to structured text.

    Returns:
        dict with keys: text, metadata, pages (one page per sheet)
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    logger.info(f"Parsing Excel: {path.name}")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)

        pages = []
        all_text_parts = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_text_parts = [f"## Sheet: {sheet_name}\n"]

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # First row as headers
            headers = [str(cell) if cell is not None else "" for cell in rows[0]]
            sheet_text_parts.append(" | ".join(headers))
            sheet_text_parts.append("-" * 40)

            for row in rows[1:]:
                cells = [str(cell) if cell is not None else "" for cell in row]
                # Also create key-value pairs for better semantic search
                row_parts = []
                for h, v in zip(headers, cells):
                    if v and h:
                        row_parts.append(f"{h}: {v}")
                    elif v:
                        row_parts.append(v)

                if row_parts:
                    sheet_text_parts.append("; ".join(row_parts))

            total_rows += len(rows) - 1  # Exclude header row
            sheet_text = "\n".join(sheet_text_parts)
            pages.append({"page": len(pages) + 1, "text": sheet_text, "sheet_name": sheet_name})
            all_text_parts.append(sheet_text)

        wb.close()
        full_text = "\n\n".join(all_text_parts)

        return {
            "text": full_text,
            "metadata": {
                "format": "excel",
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "total_rows": total_rows,
                "char_count": len(full_text),
            },
            "pages": pages,
        }

    except ImportError:
        raise ValueError("openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        logger.error(f"Excel parsing failed for {path.name}: {e}")
        raise ValueError(f"Failed to parse Excel file: {e}") from e
